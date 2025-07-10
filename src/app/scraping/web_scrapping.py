import requests
import psycopg2
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
import os

# CONEXAO COM O BANCO
load_dotenv()

def get_connection():
    return psycopg2.connect(
        dbname=os.getenv("DB_DATABASE"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        host=os.getenv("DB_HOST"),
        port=os.getenv("DB_PORT")
    )
resposta = requests.get('https://ufu.br/')
print(resposta)

if(resposta.status_code == 200):
    soup = BeautifulSoup(resposta.content, 'html.parser')
    # print(soup.prettify())
    # print(soup.title.contents)
    barra_esquerda = soup.find('ul', class_ = 'sidebar-nav nav-level-0')
    # print(barra_esquerda)
    linhas_barra_esquerda = barra_esquerda.find_all('li', class_ = 'nav-item')

    iniciar_captura = False
    linhas_desejadas_barra_esquerda = []
    links_barra_esquerda = []
    for linha in linhas_barra_esquerda:
        if 'Graduação' in linha.text.strip():
            iniciar_captura = True
        
        if iniciar_captura:
            linhas_desejadas_barra_esquerda.append(linha.text.strip())
            links_barra_esquerda.append(linha.a.get('href'))
    # print(linhas_desejadas_barra_esquerda)
    # print(links_barra_esquerda)
    # print(linha.text)
    
    for link in links_barra_esquerda:
        print("https://ufu.br"+link)

    orangeFill = PatternFill(start_color='FFC000',
                    end_color='FFC000',
                    fill_type='solid')
    
    blueFill = PatternFill(start_color='00B0F0',
                    end_color='00B0F0',
                    fill_type='solid')
    
    wb = Workbook()
    
    #openpyxl cria automaticamente um sheet chamado "Sheet" quando chamamos a função Workbook(). Por isso estou deletando.
    del wb['Sheet']
    
    sheet = wb.create_sheet('Nome do Sheet')
    
    sheet.cell(1,1).value = "Menu Nav"
    sheet.cell(1,1).fill = orangeFill
    
    sheet['B1'] = "Links"
    sheet['B1'].fill = blueFill
    
    for i, linha in enumerate(linhas_desejadas_barra_esquerda):
        sheet.cell(i+2,1).value = linha
    
    for i, link in enumerate(links_barra_esquerda):
        sheet.cell(i+2,2).value = "https://ufu.br"+link
    
    wb.save("UFU_menu_nav.xlsx")

conn = get_connection()
cursor = conn.cursor()

for menu, link in zip(linhas_desejadas_barra_esquerda, links_barra_esquerda):
    full_link = "https://ufu.br" + link
    now = datetime.now()
    cursor.execute("""
        INSERT INTO menu_nav (menu_nav, link, created_at)
        VALUES (%s, %s, %s)
    """, (menu, full_link, now))

conn.commit()
cursor.close()
conn.close()